#!/usr/bin/env python3
"""Build 2022_baseline.json from the NVI 2022 parliamentary XLSX files.

Reads:
  2022_parlamenti/Egyéni_szavazás_erjkv.xlsx    # 106 constituency results
  2022_parlamenti/Területi_listás_részerjkv.xlsx # 20 county list results

Writes:
  2022_baseline.json
    {
      "constituencies": {
        "BP-01": {
          "turnoutPct": 68.12,
          "valasztopolgar": 59754,
          "megjelentek": 45391,
          "fideszPct": 42.18,       # Fidesz-KDNP share of valid votes
          "oppPct":    48.94,       # United Opposition bloc share
          "winner":    "opposition" # or "fidesz"
        },
        ...
      },
      "counties": {
        "Budapest": {
          "turnoutPct": 75.33,
          "fideszPct": 43.22,
          "oppPct":    48.01
        },
        ...
      }
    }

Why this file exists:
  The popular-vote / county / constituency live maps show 2026 state.
  turnout-change.html uses this 2022 baseline as the x-axis (Fidesz
  stronghold strength) and y-axis (turnout delta). The baseline never
  changes, so we compute it once and commit the JSON.

Run once after checking out the repo:
  python3 build_2022_baseline.py
"""

import glob
import json
import os
import sys

import openpyxl

# Maps from the live scraper — kept in sync by hand (20 counties).
MAZ_TO_COUNTY = {
    '01': 'Budapest', '02': 'Baranya', '03': 'Bács-Kiskun', '04': 'Békés',
    '05': 'Borsod-Abaúj-Zemplén', '06': 'Csongrád-Csanád', '07': 'Fejér',
    '08': 'Győr-Moson-Sopron', '09': 'Hajdú-Bihar', '10': 'Heves',
    '11': 'Jász-Nagykun-Szolnok', '12': 'Komárom-Esztergom', '13': 'Nógrád',
    '14': 'Pest', '15': 'Somogy', '16': 'Szabolcs-Szatmár-Bereg',
    '17': 'Tolna', '18': 'Vas', '19': 'Veszprém', '20': 'Zala',
}
MAZ_TO_ID = {
    '01': 'BP', '02': 'BA', '03': 'BK', '04': 'BE', '05': 'BO', '06': 'CS',
    '07': 'FE', '08': 'GY', '09': 'HB', '10': 'HE', '11': 'JN', '12': 'KE',
    '13': 'NO', '14': 'PE', '15': 'SO', '16': 'SZ', '17': 'TO', '18': 'VA',
    '19': 'VE', '20': 'ZA',
}

# Hungarian 2022 parliamentary election — at-the-polls turnout curves.
# Originally pulled from valasztas.hu live feed via Europe Elects'
# 2022 X/Twitter coverage (national-only). NOW computed exactly from
# the official Napközbeni_részvételi_jelentések xlsx, which gives us
# per-county granularity. parse_napkozbeni() is what populates these.
#
# Used to compute a time-adjusted 2022 baseline per constituency:
#   2022-at-time-T (district) ≈ 2022-final (district)
#                                × (county_curve[T] / county_curve[18:30])
#
# Per-county curves remove the (small) urban-vs-rural intraday-shape
# bias that a single national curve baked in.
HOURLY_TIMES = ['07:00', '09:00', '11:00', '13:00', '15:00', '17:00', '18:30']

# Final including mail-in ballots; from Országos_listás_eredmény.xlsx
# (5,717,182 megjelentek / 8,215,304 eligible).
NATIONAL_FINAL_2022 = 69.59

# County name normalisation: napközbeni xlsx uses uppercase Hungarian
# names; the rest of the project uses MAZ_TO_COUNTY's mixed-case form.
NAPKOZBENI_TO_COUNTY = {
    'BUDAPEST': 'Budapest',
    'BARANYA': 'Baranya',
    'BÁCS-KISKUN': 'Bács-Kiskun',
    'BÉKÉS': 'Békés',
    'BORSOD-ABAÚJ-ZEMPLÉN': 'Borsod-Abaúj-Zemplén',
    'CSONGRÁD-CSANÁD': 'Csongrád-Csanád',
    'FEJÉR': 'Fejér',
    'GYŐR-MOSON-SOPRON': 'Győr-Moson-Sopron',
    'HAJDÚ-BIHAR': 'Hajdú-Bihar',
    'HEVES': 'Heves',
    'JÁSZ-NAGYKUN-SZOLNOK': 'Jász-Nagykun-Szolnok',
    'KOMÁROM-ESZTERGOM': 'Komárom-Esztergom',
    'NÓGRÁD': 'Nógrád',
    'PEST': 'Pest',
    'SOMOGY': 'Somogy',
    'SZABOLCS-SZATMÁR-BEREG': 'Szabolcs-Szatmár-Bereg',
    'TOLNA': 'Tolna',
    'VAS': 'Vas',
    'VESZPRÉM': 'Veszprém',
    'ZALA': 'Zala',
}


# Party-name matching for 2022. The XLSX uses the long registered
# party names; we classify each votes row as either Fidesz-KDNP, the
# United Opposition bloc, Mi Hazánk, MKKP, or "other".
def classify_2022_party(name):
    if not name:
        return 'other'
    n = name.upper()
    if 'FIDESZ' in n:
        return 'fidesz'
    # The United Opposition bloc appears as one joint ticket:
    # "DEMOKRATIKUS KOALÍCIÓ-JOBBIK MAGYARORSZÁGÉRT MOZGALOM-..."
    if 'DEMOKRATIKUS KOALÍCIÓ' in n or n.startswith('DK-') or 'JOBBIK' in n:
        return 'opp'
    if 'MI HAZÁNK' in n or 'MI HAZANK' in n:
        return 'mihazank'
    if 'KÉTFARKÚ' in n or 'KETFARKU' in n or n == 'MKKP':
        return 'mkkp'
    return 'other'


def _find_xlsx(pattern_fragments):
    """Locate a 2022_parlamenti file by matching substrings in the
    name (handles the odd Hungarian encoding in the checked-in names)."""
    all_files = glob.glob('2022_parlamenti/*.xlsx')
    for f in all_files:
        lower = f.lower()
        if all(frag in lower for frag in pattern_fragments):
            return f
    return None


def parse_constituencies(path):
    """Parse Egyéni_szavazás_erjkv.xlsx.

    Format: each constituency is one F (header) row followed by several
    T (candidate) rows. The F row has VÁLASZTÓPOLGÁR (eligible),
    MEGJELENTEK (voted), ÉRVÉNYES (valid). T rows carry candidate
    name, SZERVEZET (party), SZAVAZAT (votes), NYERTES (winner flag).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header = None
    for row in ws.iter_rows(max_row=1, values_only=True):
        header = list(row)
        break
    assert header is not None
    idx = {name: i for i, name in enumerate(header)}

    def col(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    results = {}
    current = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        tipus = col(row, 'TIPUS')
        if tipus == 'F':
            maz = col(row, 'MEGYEKÓD')
            oevk = col(row, 'OEVK')
            if maz is None or oevk is None:
                current = None
                continue
            maz = str(maz).zfill(2)
            oevk = int(str(oevk))
            if maz not in MAZ_TO_ID:
                current = None
                continue
            prefix = MAZ_TO_ID[maz]
            district_id = f'{prefix}-{oevk:02d}'
            vp = col(row, 'VÁLASZTÓPOLGÁR') or 0
            megj = col(row, 'MEGJELENTEK') or 0
            ervenyes = col(row, 'ÉRVÉNYES') or 0
            current = {
                'id': district_id,
                'vp': int(vp),
                'megj': int(megj),
                'ervenyes': int(ervenyes),
                'votes': {'fidesz': 0, 'opp': 0, 'mihazank': 0, 'mkkp': 0, 'other': 0},
                'winner': None,
            }
            results[district_id] = current
        elif tipus == 'T' and current is not None:
            szervezet = col(row, 'SZERVEZET')
            szavazat = col(row, 'SZAVAZAT') or 0
            nyertes = col(row, 'NYERTES')
            if not szervezet:
                continue
            bucket = classify_2022_party(szervezet)
            current['votes'][bucket] += int(szavazat)
            if nyertes and str(nyertes).strip():
                current['winner'] = bucket

    wb.close()

    # Compute derived percentages
    out = {}
    for did, r in results.items():
        vp = r['vp']
        megj = r['megj']
        erv = r['ervenyes'] or 1
        fidesz_pct = r['votes']['fidesz'] / erv * 100
        opp_pct = r['votes']['opp'] / erv * 100
        # Derive county from the district id prefix (e.g. "BP-01" → "Budapest")
        prefix = did.split('-')[0]
        maz = next((k for k, v in MAZ_TO_ID.items() if v == prefix), None)
        county = MAZ_TO_COUNTY.get(maz) if maz else None
        out[did] = {
            'turnoutPct': round(megj / vp * 100, 2) if vp else 0,
            'valasztopolgar': vp,
            'megjelentek': megj,
            'fideszPct': round(fidesz_pct, 2),
            'oppPct': round(opp_pct, 2),
            'winner': 'fidesz' if fidesz_pct > opp_pct else 'opposition',
            'county': county,
        }
    return out


def parse_county_list(path):
    """Parse Területi_listás_részerjkv.xlsx for per-county turnout and
    (optionally) Fidesz list share. Same F/T pattern as constituencies."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header = None
    for row in ws.iter_rows(max_row=1, values_only=True):
        header = list(row)
        break
    idx = {name: i for i, name in enumerate(header)}

    def col(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    results = {}
    current = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        tipus = col(row, 'TIPUS')
        if tipus == 'F':
            maz = col(row, 'MEGYEKÓD')
            if maz is None:
                current = None
                continue
            maz = str(maz).zfill(2)
            if maz not in MAZ_TO_COUNTY:
                current = None
                continue
            county = MAZ_TO_COUNTY[maz]
            vp = col(row, 'VÁLASZTÓPOLGÁR') or 0
            megj = col(row, 'MEGJELENTEK') or 0
            ervenyes = col(row, 'ÉRVÉNYES') or 0
            current = {
                'county': county,
                'vp': int(vp),
                'megj': int(megj),
                'ervenyes': int(ervenyes),
                'fideszVotes': 0,
                'oppVotes': 0,
            }
            results[county] = current
        elif tipus == 'T' and current is not None:
            # Territorial list file stores party list name in PÁRT_LISTA
            # (col 12) and votes in PÁRT_LISTA_SZAVAZAT (col 13).
            # Nationality lists go in NEMZ_LISTA / NEMZ_LISTA_SZAVAZAT.
            list_name = col(row, 'PÁRT_LISTA')
            votes = col(row, 'PÁRT_LISTA_SZAVAZAT') or 0
            if not list_name:
                continue
            bucket = classify_2022_party(list_name)
            if bucket == 'fidesz':
                current['fideszVotes'] += int(votes)
            elif bucket == 'opp':
                current['oppVotes'] += int(votes)

    wb.close()

    out = {}
    for county, r in results.items():
        vp = r['vp']
        megj = r['megj']
        erv = r['ervenyes'] or 1
        out[county] = {
            'turnoutPct': round(megj / vp * 100, 2) if vp else 0,
            'valasztopolgar': vp,
            'megjelentek': megj,
            'fideszPct': round(r['fideszVotes'] / erv * 100, 2),
            'oppPct': round(r['oppVotes'] / erv * 100, 2),
        }
    return out


def parse_napkozbeni(path):
    """Parse Napközbeni_részvételi_jelentések_OGY2022.xlsx.

    The file has 7 sheets, one per snapshot time (07_00, 09_00, ...,
    18_30). Each sheet has 3178 rows of municipality-level data with
    columns: Megye, Település, eligible voters, megjelentek, turnout %.

    Returns:
        {
          'national': {'07:00': pct, '09:00': pct, ...},
          'byCounty': {
            'Budapest': {'07:00': pct, ..., '18:30': pct},
            'Baranya':  {...},
            ...
          },
        }
    All values are at-the-polls (mail-in not yet counted).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    by_county = {NAPKOZBENI_TO_COUNTY[k]: {} for k in NAPKOZBENI_TO_COUNTY}
    national = {}

    for sheet_name in HOURLY_TIMES:
        # Sheet names use underscore: 07_00, 09_00, ...
        sheet_key = sheet_name.replace(':', '_')
        if sheet_key not in wb.sheetnames:
            continue
        ws = wb[sheet_key]

        # Aggregate eligible + voted per county across all rows
        county_elig = {NAPKOZBENI_TO_COUNTY[k]: 0 for k in NAPKOZBENI_TO_COUNTY}
        county_voted = {NAPKOZBENI_TO_COUNTY[k]: 0 for k in NAPKOZBENI_TO_COUNTY}
        nat_elig = 0
        nat_voted = 0

        # Skip header row 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            megye_raw = row[0]
            elig = row[2]
            voted = row[3]
            if not megye_raw or elig is None or voted is None:
                continue
            try:
                e = int(elig); v = int(voted)
            except (TypeError, ValueError):
                continue
            county = NAPKOZBENI_TO_COUNTY.get(megye_raw)
            if county:
                county_elig[county] += e
                county_voted[county] += v
            nat_elig += e
            nat_voted += v

        # Compute county-level turnout for this snapshot
        for county in by_county:
            ce = county_elig[county]
            cv = county_voted[county]
            by_county[county][sheet_name] = round(cv / ce * 100, 2) if ce else 0.0
        national[sheet_name] = round(nat_voted / nat_elig * 100, 2) if nat_elig else 0.0

    wb.close()
    return {'national': national, 'byCounty': by_county}


def parse_national_list(path):
    """Parse Országos_listás_eredmény.xlsx sheet 1 for the final 2022
    nationwide party list percentages.

    Relevant columns: PÁRT_LISTA (name), PÁRT_LISTA_SZAVAZAT (votes),
    ÉRVÉNYES (total valid). ÉRVÉNYES is on the single F-type row.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header = None
    for row in ws.iter_rows(max_row=1, values_only=True):
        header = list(row)
        break
    idx = {name: i for i, name in enumerate(header)}

    def col(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    ervenyes = 0
    votes = {'fidesz': 0, 'opp': 0, 'mihazank': 0, 'mkkp': 0}
    for row in ws.iter_rows(min_row=2, values_only=True):
        erv = col(row, 'ÉRVÉNYES')
        if erv and not ervenyes:
            try:
                ervenyes = int(erv)
            except (TypeError, ValueError):
                pass
        list_name = col(row, 'PÁRT_LISTA')
        list_votes = col(row, 'PÁRT_LISTA_SZAVAZAT')
        if list_name and list_votes:
            bucket = classify_2022_party(list_name)
            if bucket in votes:
                try:
                    votes[bucket] += int(list_votes)
                except (TypeError, ValueError):
                    pass
    wb.close()

    if not ervenyes:
        return {}
    return {
        'ervenyes': ervenyes,
        'fideszPct':    round(votes['fidesz']   / ervenyes * 100, 2),
        'oppBlocPct':   round(votes['opp']      / ervenyes * 100, 2),
        'miHazankPct':  round(votes['mihazank'] / ervenyes * 100, 2),
        'mkkpPct':      round(votes['mkkp']     / ervenyes * 100, 2),
    }


def main():
    const_path = _find_xlsx(['egy', 'erjkv']) or _find_xlsx(['egy', 'sz', 'erjkv'])
    if not const_path or 'szk' in const_path.lower():
        const_path = None
        for f in glob.glob('2022_parlamenti/*.xlsx'):
            low = f.lower()
            if 'erjkv' in low and 'szk' not in low and 'ter' not in low:
                const_path = f
                break
    county_path = None
    for f in glob.glob('2022_parlamenti/*.xlsx'):
        low = f.lower()
        if 'ter' in low and 'erjkv' in low:
            county_path = f
            break

    national_path = None
    for f in glob.glob('2022_parlamenti/*.xlsx'):
        low = f.lower()
        if 'orsz' in low and 'list' in low and 'eredm' in low:
            national_path = f
            break

    # Napközbeni hourly turnout report (per municipality, 7 sheets)
    napkozbeni_path = None
    for f in glob.glob('*.xlsx') + glob.glob('2022_parlamenti/*.xlsx'):
        if 'napk' in f.lower():
            napkozbeni_path = f
            break

    if not const_path:
        print('ERROR: could not find constituency xlsx', file=sys.stderr)
        sys.exit(1)
    if not county_path:
        print('ERROR: could not find county xlsx', file=sys.stderr)
        sys.exit(1)

    print(f'Reading {const_path}')
    constituencies = parse_constituencies(const_path)
    print(f'  parsed {len(constituencies)} constituencies')

    print(f'Reading {county_path}')
    counties = parse_county_list(county_path)
    print(f'  parsed {len(counties)} counties')

    national = {}
    if national_path:
        print(f'Reading {national_path}')
        national = parse_national_list(national_path)
        print(f'  national list: {national}')

    hourly_data = {'national': {}, 'byCounty': {}}
    if napkozbeni_path:
        print(f'Reading {napkozbeni_path}')
        hourly_data = parse_napkozbeni(napkozbeni_path)
        print(f'  national curve: {hourly_data["national"]}')
        print(f'  parsed curves for {len(hourly_data["byCounty"])} counties')

    out = {
        'generatedFrom': 'NVI 2022 parliamentary XLSX protocol files',
        'constituencies': constituencies,
        'counties': counties,
        'nationalVote2022': national,
        'hourlyNational2022': hourly_data['national'],
        'hourlyByCounty2022': hourly_data['byCounty'],
        'nationalFinal2022': NATIONAL_FINAL_2022,
    }

    with open('2022_baseline.json', 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2, sort_keys=True)
    print('Wrote 2022_baseline.json')

    # Quick sanity summary
    if constituencies:
        ts = [v['turnoutPct'] for v in constituencies.values()]
        fs = [v['fideszPct'] for v in constituencies.values()]
        print(f'  constituency turnout range: {min(ts):.1f}% – {max(ts):.1f}%')
        print(f'  constituency Fidesz range:  {min(fs):.1f}% – {max(fs):.1f}%')
        print(f'  sample: BP-01 = {constituencies.get("BP-01")}')
    if counties:
        print(f'  sample: Budapest = {counties.get("Budapest")}')


if __name__ == '__main__':
    main()
